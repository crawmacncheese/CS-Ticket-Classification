"""Build multi-sheet Excel workbook for portal download (tickets + tier breakdown)."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill

from cs_tickets.pipeline import format_cell
from cs_tickets.portal_stats import tier_stats_sheet_rows
from cs_tickets.run_metadata import RunMetadata
from cs_tickets.schema import MASTER_COLUMNS

CELL_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)

_HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_GRAND_FILL = PatternFill("solid", fgColor="D4DCE8")
_ZEBRA_A = PatternFill("solid", fgColor="E8EEF6")
_ZEBRA_B = PatternFill("solid", fgColor="F5F8FC")


def _apply_cell_style(
    cell: Cell,
    *,
    fill: PatternFill | None = None,
    font: Font | None = None,
) -> None:
    cell.alignment = CELL_ALIGN
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font


def _write_metadata_sheet(wb: Workbook, metadata: RunMetadata) -> None:
    ws = wb.create_sheet("Run metadata", 0)
    for col, title in enumerate(("Field", "Value"), start=1):
        _apply_cell_style(ws.cell(1, col, title), font=Font(bold=True))
    for row_idx, (field, value) in enumerate(metadata.metadata_sheet_rows(), start=2):
        _apply_cell_style(ws.cell(row_idx, 1, field))
        _apply_cell_style(ws.cell(row_idx, 2, value))


def build_run_workbook_bytes(
    rows: list[dict[str, Any]],
    metadata: RunMetadata | None = None,
) -> bytes:
    wb = Workbook()
    ws_t = wb.active
    assert ws_t is not None
    ws_t.title = "Tickets"
    for col, name in enumerate(MASTER_COLUMNS, start=1):
        _apply_cell_style(ws_t.cell(1, col, name), font=Font(bold=True))
    for ri, row in enumerate(rows, start=2):
        for ci, key in enumerate(MASTER_COLUMNS, start=1):
            _apply_cell_style(ws_t.cell(ri, ci, format_cell(row.get(key))))

    ws_s = wb.create_sheet("Tier breakdown")
    header, data = tier_stats_sheet_rows(rows)
    for col, h in enumerate(header, start=1):
        _apply_cell_style(
            ws_s.cell(1, col, h),
            fill=_HEADER_FILL,
            font=_HEADER_FONT,
        )
    for i, r in enumerate(data):
        ridx = i + 2
        is_grand = r[0] == "Grand Total"
        zebra = _ZEBRA_A if i % 2 == 0 else _ZEBRA_B
        for col, val in enumerate(r, start=1):
            if is_grand:
                _apply_cell_style(
                    ws_s.cell(ridx, col, val),
                    fill=_GRAND_FILL,
                    font=Font(bold=True),
                )
            else:
                _apply_cell_style(ws_s.cell(ridx, col, val), fill=zebra)

    if metadata is not None:
        _write_metadata_sheet(wb, metadata)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
