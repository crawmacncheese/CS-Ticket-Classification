from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from cs_tickets.feedback.ids import normalize_ticket_id
from cs_tickets.feedback.mine_rules import count_already_classified, mine_rule_proposals
from cs_tickets.feedback.mine_taxonomy import mine_taxonomy_proposals
from cs_tickets.feedback.models import MineConfig, RuleProposal, TaxonomyProposal
from cs_tickets.classifier_rules import RuleSpec
from cs_tickets.schema import TIER_COLUMNS
from cs_tickets.taxonomy import AllowList, resolve_classified_upload_sheet

_REFERENCE_SHEET = "SCMP_Tickets_Master_Categorized"
LEARN_SHEET = _REFERENCE_SHEET  # backward compat for tests
_SIGNAL_COLUMNS = ("tags", "subject", "raw_subject", "description")
_REQUIRED_TIER_COLUMNS = TIER_COLUMNS[:4]


@dataclass(frozen=True)
class LearnParseResult:
    upload_id: str
    filename: str
    row_count: int
    eligible_row_count: int
    distinct_tier_paths: int
    rule_proposal_count: int
    taxonomy_proposal_count: int
    rule_proposals: tuple[RuleProposal, ...] = ()
    taxonomy_proposals: tuple[TaxonomyProposal, ...] = ()
    already_classified_count: int = 0

    @property
    def status(self) -> str:
        return "processed"


def _tier_tuple(row: dict[str, object]) -> tuple[str, str, str, str, str]:
    parts = [str(row.get(col) or "").strip() for col in TIER_COLUMNS[:4]]
    granular = str(row.get("Granular_Tech_UI_Type") or "").strip() or "N/A"
    return (*parts, granular)


def _has_signal(row: dict[str, object]) -> bool:
    return any(str(row.get(col) or "").strip() for col in _SIGNAL_COLUMNS)


def _row_is_eligible(row: dict[str, object]) -> bool:
    if row.get("id") is None or str(row.get("id") or "").strip() == "":
        return False
    if not any(_tier_tuple(row)):
        return False
    return _has_signal(row)


def _read_workbook_rows(path: Path) -> list[dict[str, object]]:
    sheet = resolve_classified_upload_sheet(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet!r} not found in workbook")
        ws = wb[sheet]
        rows_iter = ws.iter_rows(min_row=1, values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            raise ValueError(f"Sheet {sheet!r} is empty")
        header = [str(h).strip() if h is not None else "" for h in header_row]
        idx = {name: i for i, name in enumerate(header) if name}

        missing = [col for col in ("id", *_REQUIRED_TIER_COLUMNS) if col not in idx]
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(missing)}")
        if not any(col in idx for col in _SIGNAL_COLUMNS):
            raise ValueError("Workbook must include at least one of: tags, subject, raw_subject, description")

        parsed: list[dict[str, object]] = []
        for row in rows_iter:
            if not row or row[0] is None:
                continue
            record = {
                col: row[idx[col]] if idx[col] < len(row) else None
                for col in idx
            }
            if "id" in record:
                record["id"] = normalize_ticket_id(record["id"]) or None
            if "Granular_Tech_UI_Type" not in record:
                record["Granular_Tech_UI_Type"] = "N/A"
            parsed.append(record)
        return parsed
    finally:
        wb.close()


def parse_categorized_workbook(
    path: Path,
    *,
    upload_id: str,
    filename: str,
    allow: AllowList | None = None,
    mine: bool = True,
    mine_config: MineConfig | None = None,
    existing_rules: tuple[RuleSpec, ...] | None = None,
) -> LearnParseResult:
    """Read classified upload workbook, optionally mine rule/taxonomy proposals."""
    parsed = _read_workbook_rows(path)
    eligible = [r for r in parsed if _row_is_eligible(r)]
    tier_paths = {_tier_tuple(r) for r in eligible if any(_tier_tuple(r))}

    rule_proposals: tuple[RuleProposal, ...] = ()
    taxonomy_proposals: tuple[TaxonomyProposal, ...] = ()
    already_classified = 0

    if mine and allow is not None:
        cfg = mine_config or MineConfig()
        rule_proposals = mine_rule_proposals(
            eligible, allow, config=cfg, existing_rules=existing_rules
        )
        taxonomy_proposals = mine_taxonomy_proposals(eligible, allow, config=cfg)
        already_classified = count_already_classified(eligible, allow)

    return LearnParseResult(
        upload_id=upload_id,
        filename=filename,
        row_count=len(parsed),
        eligible_row_count=len(eligible),
        distinct_tier_paths=len(tier_paths),
        rule_proposal_count=len(rule_proposals),
        taxonomy_proposal_count=len(taxonomy_proposals),
        rule_proposals=rule_proposals,
        taxonomy_proposals=taxonomy_proposals,
        already_classified_count=already_classified,
    )
