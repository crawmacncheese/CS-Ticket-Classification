from __future__ import annotations

import csv
from collections import Counter
from collections.abc import Iterator
from dataclasses import dataclass
from difflib import get_close_matches
from pathlib import Path

from openpyxl import load_workbook

from cs_tickets.schema import MASTER_COLUMNS, PIPELINE_FALLBACK_TIER_TUPLES, TIER_COLUMNS

_GRANULAR_COLUMN = "Granular_Tech_UI_Type"
_GRANULAR_DEFAULT = "N/A"
_REQUIRED_TIER_COLUMNS = TIER_COLUMNS[:4]
_REFERENCE_SHEET = "SCMP_Tickets_Master_Categorized"
_PORTAL_SHEET = "Tickets"


@dataclass(frozen=True)
class AllowList:
    """Valid (Tier1..Tier5) tuples for classifier output validation."""
    tuples: frozenset[tuple[str, str, str, str, str]]

    def __contains__(self, item: object) -> bool:
        return item in self.tuples


@dataclass(frozen=True)
class AllowlistRejectionInfo:
    """Why a suggested 5-tuple was rejected from the allow-list."""

    cause: str  # hallucinated | typo_close_match | taxonomy_not_allowlisted
    message: str
    rejected_tier: tuple[str, str, str, str, str] | None
    close_match_tier: tuple[str, str, str, str, str] | None
    can_add_to_allowlist: bool
    novelty_type: str | None = None

    @property
    def close_match_path(self) -> str:
        if not self.close_match_tier:
            return ""
        return " → ".join(self.close_match_tier[:4])

    @property
    def rejected_path(self) -> str:
        if not self.rejected_tier:
            return ""
        return " → ".join(self.rejected_tier[:4])


def load_taxonomy_four_tuples(taxonomy_csv: Path | None) -> frozenset[tuple[str, str, str, str]]:
    if not taxonomy_csv or not taxonomy_csv.is_file():
        return frozenset()
    return frozenset(five[:4] for five, _ in iter_taxonomy_pivot_rows(taxonomy_csv))


def _tier_path_display(tier: tuple[str, str, str, str, str]) -> str:
    return " → ".join(tier[:4])


def _find_close_allowlist_match(
    candidate: tuple[str, str, str, str, str],
    allow: AllowList,
) -> tuple[str, str, str, str, str] | None:
    """Near-miss allow-list tuple (typo / wrong granular), not a taxonomy gap."""
    t1, t2, t3, t4, _granular = candidate
    same4 = [t for t in allow.tuples if t[:4] == candidate[:4]]
    if same4:
        return sorted(same4)[0]

    same3 = [t for t in allow.tuples if t[:3] == (t1, t2, t3)]
    if not same3:
        return None
    t4_options = sorted({t[3] for t in same3})
    if t4 in t4_options:
        return None
    close = get_close_matches(t4, t4_options, n=1, cutoff=0.8)
    if not close:
        return None
    return next(t for t in same3 if t[3] == close[0])


def classify_allowlist_rejection(
    candidate: tuple[str, str, str, str, str],
    allow: AllowList,
    taxonomy_csv: Path | None,
) -> AllowlistRejectionInfo:
    """Classify why a model-suggested tier is outside the allow-list."""
    if candidate in allow.tuples:
        return AllowlistRejectionInfo(
            cause="already_allowlisted",
            message="Category is already in the allow-list.",
            rejected_tier=candidate,
            close_match_tier=None,
            can_add_to_allowlist=False,
        )

    rejected_path = _tier_path_display(candidate)
    taxonomy_four = load_taxonomy_four_tuples(taxonomy_csv)
    novelty = novelty_type_for_tuple(candidate, allow)
    close = _find_close_allowlist_match(candidate, allow)

    in_taxonomy = candidate[:4] in taxonomy_four
    can_add = novelty == "granular_new" or (in_taxonomy and novelty != TIER1_NOVELTY)

    if close and novelty != "granular_new":
        return AllowlistRejectionInfo(
            cause="typo_close_match",
            message=(
                f"Close allow-list match: {_tier_path_display(close)} "
                f"(model suggested: {rejected_path})."
            ),
            rejected_tier=candidate,
            close_match_tier=close,
            can_add_to_allowlist=False,
            novelty_type=novelty,
        )

    if can_add:
        return AllowlistRejectionInfo(
            cause="taxonomy_not_allowlisted",
            message=(
                f"Category is in the taxonomy but not yet in the allow-list "
                f"(no workbook exemplar): {rejected_path}."
            ),
            rejected_tier=candidate,
            close_match_tier=close,
            can_add_to_allowlist=True,
            novelty_type=novelty,
        )

    return AllowlistRejectionInfo(
        cause="hallucinated",
        message=(
            f"Suggested category is not in the taxonomy or allow-list: {rejected_path}."
        ),
        rejected_tier=candidate,
        close_match_tier=close,
        can_add_to_allowlist=False,
        novelty_type=novelty,
    )


def _parse_pivot_taxonomy_csv(path: Path) -> set[tuple[str, str, str, str]]:
    return {five[:4] for five, _ in iter_taxonomy_pivot_rows(path)}


def iter_taxonomy_pivot_rows(
    path: Path,
) -> list[tuple[tuple[str, str, str, str, str], str]]:
    """Parse pivot-style CSV.

    Returns list of (tier1..tier4, granular_N/A), count_cell) for each data row
    that defines a leaf/summary path in the pivot (same rows that contribute to the allow-list).
    """
    text = path.read_text(encoding="utf-8-sig", errors="replace").splitlines()
    reader = csv.reader(text)
    header = next(reader, None)
    if not header or len(header) < 4:
        return []
    out: list[tuple[tuple[str, str, str, str, str], str]] = []
    carry = ["", "", "", ""]
    for row in reader:
        if not row or all(not (c or "").strip() for c in row[:4]):
            continue
        raw = [(row[i] if i < len(row) else "") or "" for i in range(4)]
        joined = "".join(raw).strip()
        if joined.lower().startswith("grand total"):
            break
        for i in range(4):
            cell = raw[i].strip()
            if cell:
                carry[i] = cell
                for j in range(i + 1, 4):
                    carry[j] = ""
        t1, t2, t3, t4 = carry
        count_cell = row[4].strip() if len(row) > 4 else ""
        if not t1:
            continue
        if t2 == "Junk" and not t3 and not t4:
            five = (t1, "Junk", "Junk", "Junk", "N/A")
            out.append((five, count_cell))
            continue
        if t2 and t3 and t4:
            five = (t1, t2, t3, t4, "N/A")
            out.append((five, count_cell))
    return out


def load_taxonomy_raw_grid(path: Path, *, max_body_rows: int = 600) -> tuple[list[str], list[list[str]]]:
    """Return trimmed header + body rows from Taxonomy.csv (for UI), up to Grand Total inclusive."""
    text = path.read_text(encoding="utf-8-sig", errors="replace").splitlines()
    reader = csv.reader(text)
    header = next(reader, None)
    if not header:
        return [], []

    def trim(cells: list[str]) -> list[str]:
        out = list(cells)
        while out and not (out[-1] or "").strip():
            out.pop()
        return out

    header = trim(list(header))
    body: list[list[str]] = []
    for row in reader:
        cells = trim(list(row))
        if not cells:
            continue
        joined = "".join(cells[:4]).strip().lower()
        if joined.startswith("grand total"):
            body.append(cells[: max(len(header), len(cells))])
            break
        body.append(cells)
        if len(body) >= max_body_rows:
            break
    return header, body


def _load_workbook_tuples(xlsx: Path, sheet: str = "SCMP_Tickets_Master_Categorized") -> set[tuple[str, str, str, str, str]]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet {sheet!r} not in {xlsx}")
    ws = wb[sheet]
    rows = ws.iter_rows(min_row=1, values_only=True)
    header = next(rows, None)
    if not header:
        wb.close()
        return set()
    idx = {h: i for i, h in enumerate(header) if h}
    keys = list(TIER_COLUMNS)
    id_idx = idx.get("id")
    out: set[tuple[str, str, str, str, str]] = set()
    for row in rows:
        if not row:
            continue
        if id_idx is not None:
            id_val = row[id_idx] if id_idx < len(row) else None
            if id_val is None or not str(id_val).strip():
                continue
        elif row[0] is None:
            continue
        tup = _tuple_from_row_cells(row, idx)
        if any(tup):
            out.add(tuple(str(x) if x is not None else "" for x in tup))
    wb.close()
    return out


def load_allowlist(
    taxonomy_csv: Path | None,
    workbook_xlsx: Path | None,
) -> AllowList:
    """Union of 5-tuples from workbook, CSV-derived rows (Granular N/A), and pipeline fallbacks."""
    five: set[tuple[str, str, str, str, str]] = set()
    if workbook_xlsx and workbook_xlsx.is_file():
        five |= _load_workbook_tuples(workbook_xlsx)
    if taxonomy_csv and taxonomy_csv.is_file():
        for t4 in _parse_pivot_taxonomy_csv(taxonomy_csv):
            five.add((*t4, "N/A"))
    five |= PIPELINE_FALLBACK_TIER_TUPLES
    if not five:
        raise ValueError("Allow-list is empty: provide doc/Taxonomy.csv and/or doc/CS_ticket_new_categorizations.xlsx")
    return AllowList(frozenset(five))


TIER1_NOVELTY = "tier1_new"


def novelty_type_for_tuple(
    tier: tuple[str, str, str, str, str],
    allow: AllowList,
) -> str:
    """Classify how a 5-tuple differs from the current allow-list (for hybrid promote routing)."""
    t1, t2, t3, t4, granular = tier
    tier4_paths = {t[:4] for t in allow.tuples if t[0] == t1 and t[1] == t2 and t[2] == t3}
    tier3_paths = {t[:3] for t in allow.tuples if t[0] == t1 and t[1] == t2}
    tier1_set = {t[0] for t in allow.tuples}
    if t1 and t1 not in tier1_set:
        return TIER1_NOVELTY
    if (t1, t2, t3, t4) not in tier4_paths and t4:
        return "tier4_new"
    if (t1, t2, t3) not in tier3_paths and t3:
        return "tier3_new"
    if granular != "N/A":
        return "granular_new"
    return "path_new"


def split_taxonomy_proposals(
    proposals: tuple["TaxonomyProposal", ...],
) -> tuple[tuple["TaxonomyProposal", ...], frozenset[tuple[str, str, str, str, str]]]:
    """Route accepted taxonomy proposals to CSV merge vs workbook exemplar merge."""
    from cs_tickets.feedback.models import TaxonomyProposal

    csv_proposals: list[TaxonomyProposal] = []
    workbook_tuples: set[tuple[str, str, str, str, str]] = set()
    for proposal in proposals:
        novelty = proposal.novelty_type
        if novelty == "granular_new":
            workbook_tuples.add(proposal.tier)
        elif novelty in ("tier4_new", "tier3_new", TIER1_NOVELTY):
            csv_proposals.append(proposal)
            if proposal.tier[4] != "N/A":
                workbook_tuples.add(proposal.tier)
        elif novelty == "path_new":
            csv_proposals.append(proposal)
            if proposal.tier[4] != "N/A":
                workbook_tuples.add(proposal.tier)
        else:
            csv_proposals.append(proposal)
    return tuple(csv_proposals), frozenset(workbook_tuples)


def _is_complete_five_tuple(t: tuple[str, str, str, str, str]) -> bool:
    return all((v or "").strip() for v in t)


def _row_five_tuple(row: dict[str, str]) -> tuple[str, str, str, str, str]:
    vals: list[str] = []
    for col in TIER_COLUMNS:
        v = str(row.get(col) or "").strip()
        if col == _GRANULAR_COLUMN and not v:
            v = _GRANULAR_DEFAULT
        vals.append(v)
    return tuple(vals)


def _workbook_header_index(header: tuple) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        name = str(h).strip()
        if name:
            out[name] = i
    return out


def _tier_cell_value(row: tuple, idx: dict[str, int], col: str) -> str:
    if col not in idx:
        if col == _GRANULAR_COLUMN:
            return _GRANULAR_DEFAULT
        return ""
    val = row[idx[col]] if idx[col] < len(row) else None
    text = str(val).strip() if val is not None else ""
    if col == _GRANULAR_COLUMN and not text:
        return _GRANULAR_DEFAULT
    return text


def _tuple_from_row_cells(row: tuple, idx: dict[str, int]) -> tuple[str, str, str, str, str]:
    return tuple(_tier_cell_value(row, idx, col) for col in TIER_COLUMNS)


def _try_workbook_sheet_index(header: tuple) -> dict[str, int] | None:
    idx = _workbook_header_index(header)
    if any(c not in idx for c in _REQUIRED_TIER_COLUMNS):
        return None
    return idx


def _count_ticket_rows(ws, idx: dict[str, int]) -> int:
    id_idx = idx.get("id")
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        if id_idx is not None:
            id_val = row[id_idx] if id_idx < len(row) else None
            if id_val is None or not str(id_val).strip():
                continue
        elif row[0] is None:
            continue
        count += 1
    return count


def resolve_classified_upload_sheet(xlsx: Path) -> str:
    """Pick the ticket sheet in an analyst classified upload."""
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    try:
        names = wb.sheetnames
        if _REFERENCE_SHEET in names:
            return _REFERENCE_SHEET
        if _PORTAL_SHEET in names:
            return _PORTAL_SHEET
        best_name, best_count = "", -1
        for name in names:
            ws = wb[name]
            rows = ws.iter_rows(min_row=1, values_only=True)
            header = next(rows, None)
            if not header:
                continue
            idx = _try_workbook_sheet_index(header)
            if idx is None:
                continue
            count = _count_ticket_rows(ws, idx)
            if count > best_count:
                best_name, best_count = name, count
        if best_count <= 0:
            raise ValueError(f"No sheet with ticket rows and tier columns in {xlsx.name}")
        return best_name
    finally:
        wb.close()


def extract_classified_workbook_five_tuples(
    xlsx: Path,
) -> frozenset[tuple[str, str, str, str, str]]:
    sheet = resolve_classified_upload_sheet(xlsx)
    return extract_workbook_five_tuples(xlsx, sheet=sheet)


def count_classified_tickets_per_tuple(
    xlsx: Path,
) -> dict[tuple[str, str, str, str, str], int]:
    sheet = resolve_classified_upload_sheet(xlsx)
    return count_tickets_per_tuple(xlsx, sheet=sheet)


def _validate_workbook_sheet(xlsx: Path, sheet: str) -> dict[str, int]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet!r} not in {xlsx.name}")
        ws = wb[sheet]
        rows = ws.iter_rows(min_row=1, values_only=True)
        header = next(rows, None)
        if not header:
            raise ValueError(f"Sheet {sheet!r} in {xlsx.name} has no header row")
        idx = _workbook_header_index(header)
        missing = [c for c in _REQUIRED_TIER_COLUMNS if c not in idx]
        if missing:
            raise ValueError(
                f"Sheet {sheet!r} in {xlsx.name} missing tier columns: {', '.join(missing)}"
            )
        return idx
    finally:
        wb.close()


def iter_workbook_master_rows(
    xlsx: Path,
    sheet: str = "SCMP_Tickets_Master_Categorized",
) -> Iterator[dict[str, str]]:
    """Yield full master-row dicts in sheet order (header mapped via MASTER_COLUMNS)."""
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet!r} not in {xlsx.name}")
        ws = wb[sheet]
        rows = ws.iter_rows(min_row=1, values_only=True)
        header = next(rows, None)
        if not header:
            return
        idx = _workbook_header_index(header)
        id_idx = idx.get("id")
        for row in rows:
            if not row:
                continue
            if id_idx is not None:
                id_val = row[id_idx] if id_idx < len(row) else None
                if id_val is None or not str(id_val).strip():
                    continue
            elif row[0] is None:
                continue
            out: dict[str, str] = {}
            for col in MASTER_COLUMNS:
                if col in idx:
                    val = row[idx[col]]
                    out[col] = str(val) if val is not None else ""
                elif col == _GRANULAR_COLUMN:
                    out[col] = _GRANULAR_DEFAULT
                else:
                    out[col] = ""
            yield out
    finally:
        wb.close()


def extract_workbook_five_tuples(
    xlsx: Path,
    sheet: str = "SCMP_Tickets_Master_Categorized",
) -> frozenset[tuple[str, str, str, str, str]]:
    """Distinct complete 5-tuples from upload; skips incomplete rows."""
    _validate_workbook_sheet(xlsx, sheet)
    out: set[tuple[str, str, str, str, str]] = set()
    for row in iter_workbook_master_rows(xlsx, sheet=sheet):
        tup = _row_five_tuple(row)
        if _is_complete_five_tuple(tup):
            out.add(tup)
    return frozenset(out)


def count_tickets_per_tuple(
    xlsx: Path,
    sheet: str = "SCMP_Tickets_Master_Categorized",
) -> dict[tuple[str, str, str, str, str], int]:
    """Ticket counts per complete 5-tuple (for checklist UI)."""
    _validate_workbook_sheet(xlsx, sheet)
    counts: Counter[tuple[str, str, str, str, str]] = Counter()
    for row in iter_workbook_master_rows(xlsx, sheet=sheet):
        tup = _row_five_tuple(row)
        if _is_complete_five_tuple(tup):
            counts[tup] += 1
    return dict(counts)


def diff_against_allowlist(
    upload_tuples: frozenset[tuple[str, str, str, str, str]],
    allow: AllowList,
) -> frozenset[tuple[str, str, str, str, str]]:
    return frozenset(t for t in upload_tuples if t not in allow)


def merge_tuples_into_workbook(
    target_xlsx: Path,
    source_xlsx: Path,
    new_tuples: frozenset[tuple[str, str, str, str, str]],
    *,
    sheet: str = "SCMP_Tickets_Master_Categorized",
) -> int:
    """Append one exemplar row per new tuple. Return rows added."""
    if not new_tuples:
        return 0
    source_sheet = resolve_classified_upload_sheet(source_xlsx)
    _validate_workbook_sheet(target_xlsx, sheet)
    wb = load_workbook(target_xlsx, read_only=False, data_only=False)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet!r} not in {target_xlsx.name}")
        ws = wb[sheet]
        header = [cell.value for cell in ws[1]]
        idx = _workbook_header_index(tuple(header))
        missing = [c for c in MASTER_COLUMNS if c not in idx]
        if missing:
            raise ValueError(
                f"Sheet {sheet!r} in {target_xlsx.name} missing columns: {', '.join(missing)}"
            )
        exemplars: dict[tuple[str, str, str, str, str], dict[str, str]] = {}
        for row in iter_workbook_master_rows(source_xlsx, sheet=source_sheet):
            tup = _row_five_tuple(row)
            if tup in new_tuples and tup not in exemplars:
                exemplars[tup] = row
            if len(exemplars) == len(new_tuples):
                break
        added = 0
        for tup in sorted(new_tuples):
            exemplar = exemplars.get(tup)
            if not exemplar:
                continue
            values = [exemplar.get(col, "") for col in MASTER_COLUMNS]
            ws.append(values)
            added += 1
        wb.save(target_xlsx)
        return added
    finally:
        wb.close()


def append_exemplar_row(
    target_xlsx: Path,
    exemplar: dict[str, str],
    *,
    sheet: str = "SCMP_Tickets_Master_Categorized",
) -> int:
    """Append one classified exemplar row to the reference workbook."""
    _validate_workbook_sheet(target_xlsx, sheet)
    wb = load_workbook(target_xlsx, read_only=False, data_only=False)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet!r} not in {target_xlsx.name}")
        ws = wb[sheet]
        header = [cell.value for cell in ws[1]]
        idx = _workbook_header_index(tuple(header))
        missing = [c for c in MASTER_COLUMNS if c not in idx]
        if missing:
            raise ValueError(
                f"Sheet {sheet!r} in {target_xlsx.name} missing columns: {', '.join(missing)}"
            )
        ws.append([exemplar.get(col, "") for col in MASTER_COLUMNS])
        wb.save(target_xlsx)
        return 1
    finally:
        wb.close()


def resolve_exemplars_for_tuples(
    source_xlsx: Path,
    tuples: frozenset[tuple[str, str, str, str, str]],
) -> dict[tuple[str, str, str, str, str], dict[str, str]]:
    """First exemplar row per tuple from classified upload (same order as merge)."""
    if not tuples:
        return {}
    source_sheet = resolve_classified_upload_sheet(source_xlsx)
    exemplars: dict[tuple[str, str, str, str, str], dict[str, str]] = {}
    for row in iter_workbook_master_rows(source_xlsx, sheet=source_sheet):
        tup = _row_five_tuple(row)
        if tup in tuples and tup not in exemplars:
            exemplars[tup] = row
        if len(exemplars) == len(tuples):
            break
    return exemplars
