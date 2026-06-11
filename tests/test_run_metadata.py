from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from cs_tickets.portal_workbook import build_run_workbook_bytes
from cs_tickets.run_metadata import build_run_metadata, count_tbc_rows


def test_count_tbc_rows() -> None:
    rows = [
        {"Tier4_Type": "TBC (Manual Review)"},
        {"Tier4_Type": "Rate or Renewal Inquiry"},
    ]
    assert count_tbc_rows(rows) == 1


def test_workbook_includes_run_metadata_sheet() -> None:
    rows = [{"id": 1, "Tier4_Type": "N/A"}]
    meta = build_run_metadata(
        run_id="run-uuid",
        source_filename="sample.ndjson",
        rows=rows,
        warning_count=0,
    )
    payload = build_run_workbook_bytes(rows, metadata=meta)
    wb = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    assert wb.sheetnames[0] == "Run metadata"
    assert wb.sheetnames[1:] == ["Tickets", "Tier breakdown"]
    ws = wb["Run metadata"]
    assert ws["A2"].value == "run_id"
    assert ws["B2"].value == "run-uuid"
