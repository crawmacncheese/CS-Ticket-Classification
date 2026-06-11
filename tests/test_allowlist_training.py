from __future__ import annotations

import shutil
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from cs_tickets.schema import MASTER_COLUMNS, TIER_COLUMNS
from cs_tickets.taxonomy import (
    diff_against_allowlist,
    extract_classified_workbook_five_tuples,
    extract_workbook_five_tuples,
    load_allowlist,
    merge_tuples_into_workbook,
    count_tickets_per_tuple,
    resolve_classified_upload_sheet,
    _is_complete_five_tuple,
)


SHEET = "SCMP_Tickets_Master_Categorized"


def _write_workbook(path: Path, rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(list(MASTER_COLUMNS))
    for row in rows:
        ws.append([row.get(c, "") for c in MASTER_COLUMNS])
    wb.save(path)


def _sample_row(**tier_overrides: str) -> dict[str, str]:
    row = {c: "" for c in MASTER_COLUMNS}
    row["id"] = tier_overrides.pop("id", "1")
    row["subject"] = tier_overrides.pop("subject", "Test ticket")
    for col in TIER_COLUMNS:
        row[col] = tier_overrides.get(col, "Fill")
    return row


def test_is_complete_five_tuple() -> None:
    assert _is_complete_five_tuple(("A", "B", "C", "D", "E"))
    assert not _is_complete_five_tuple(("A", "B", "", "D", "E"))
    assert not _is_complete_five_tuple(("A", "B", "  ", "D", "E"))


def test_extract_skips_incomplete_rows(tmp_path: Path) -> None:
    xlsx = tmp_path / "upload.xlsx"
    complete = _sample_row(
        id="1",
        Tier1_Segment="B2C",
        Tier2_Stream="Service Task",
        Tier3_Cat="General Support",
        Tier4_Type="Account",
        Granular_Tech_UI_Type="Login",
    )
    incomplete = _sample_row(id="2", Tier3_Cat="")
    _write_workbook(xlsx, [complete, incomplete])
    tuples = extract_workbook_five_tuples(xlsx)
    assert len(tuples) == 1
    assert ("B2C", "Service Task", "General Support", "Account", "Login") in tuples


def test_extract_defaults_missing_granular_column(tmp_path: Path) -> None:
    """Analyst uploads often have Tier1–4 only; granular defaults to N/A."""
    xlsx = tmp_path / "four_tier.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    cols = list(MASTER_COLUMNS[:20]) + list(TIER_COLUMNS[:4])
    ws.append(cols)
    ws.append(
        ["http://example.com", "42", *([""] * 18), "B2C", "Service Task", "General Support", "Account"]
    )
    wb.save(xlsx)
    tuples = extract_workbook_five_tuples(xlsx)
    assert ("B2C", "Service Task", "General Support", "Account", "N/A") in tuples


def test_example_workbook_without_granular_column(repo_root: Path) -> None:
    sample = repo_root / "20260528_-_CS_ticket_new_categorizations.xlsx"
    if not sample.is_file():
        pytest.skip("example workbook missing")
    tuples = extract_workbook_five_tuples(sample)
    assert len(tuples) >= 1


def test_count_tickets_per_tuple(tmp_path: Path) -> None:
    tup = ("B2C", "Service Task", "General Support", "Account", "Login")
    row = _sample_row(
        Tier1_Segment=tup[0],
        Tier2_Stream=tup[1],
        Tier3_Cat=tup[2],
        Tier4_Type=tup[3],
        Granular_Tech_UI_Type=tup[4],
    )
    xlsx = tmp_path / "upload.xlsx"
    _write_workbook(xlsx, [row, dict(row, id="2")])
    counts = count_tickets_per_tuple(xlsx)
    assert counts[tup] == 2


def test_diff_and_merge_against_allowlist(repo_root: Path, tmp_path: Path) -> None:
    tax = repo_root / "doc" / "Taxonomy.csv"
    ref = repo_root / "doc" / "CS_ticket_new_categorizations.xlsx"
    if not tax.is_file() or not ref.is_file():
        pytest.skip("doc artifacts missing")

    allow_before = load_allowlist(tax, ref)
    novel = (
        "NovelSeg",
        "NovelStream",
        "NovelCat",
        "NovelType",
        "NovelGranular",
    )
    if novel in allow_before:
        pytest.skip("fixture tuple already in allow-list")

    upload = tmp_path / "upload.xlsx"
    first = _sample_row(
        id="100",
        subject="First exemplar",
        Tier1_Segment=novel[0],
        Tier2_Stream=novel[1],
        Tier3_Cat=novel[2],
        Tier4_Type=novel[3],
        Granular_Tech_UI_Type=novel[4],
    )
    second = _sample_row(
        id="101",
        subject="Second row same tuple",
        Tier1_Segment=novel[0],
        Tier2_Stream=novel[1],
        Tier3_Cat=novel[2],
        Tier4_Type=novel[3],
        Granular_Tech_UI_Type=novel[4],
    )
    _write_workbook(upload, [first, second])

    upload_tuples = extract_workbook_five_tuples(upload)
    new_tuples = diff_against_allowlist(upload_tuples, allow_before)
    assert novel in new_tuples

    target = tmp_path / "target.xlsx"
    shutil.copy2(ref, target)
    added = merge_tuples_into_workbook(target, upload, frozenset({novel}))
    assert added == 1

    allow_after = load_allowlist(tax, target)
    assert novel in allow_after
    assert len(allow_after.tuples) == len(allow_before.tuples) + 1


def test_merge_uses_first_exemplar_in_sheet_order(tmp_path: Path) -> None:
    novel = ("X1", "X2", "X3", "X4", "X5")
    first = _sample_row(id="first-id", subject="First", Tier1_Segment=novel[0], Tier2_Stream=novel[1], Tier3_Cat=novel[2], Tier4_Type=novel[3], Granular_Tech_UI_Type=novel[4])
    second = _sample_row(id="second-id", subject="Second", Tier1_Segment=novel[0], Tier2_Stream=novel[1], Tier3_Cat=novel[2], Tier4_Type=novel[3], Granular_Tech_UI_Type=novel[4])
    source = tmp_path / "source.xlsx"
    _write_workbook(source, [first, second])

    target = tmp_path / "target.xlsx"
    _write_workbook(target, [_sample_row(id="existing")])
    merge_tuples_into_workbook(target, source, frozenset({novel}))

    from openpyxl import load_workbook

    wb = load_workbook(target, read_only=True, data_only=True)
    ws = wb[SHEET]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    assert len(rows) == 2
    idx = {h: i for i, h in enumerate(MASTER_COLUMNS)}
    appended = rows[-1]
    assert str(appended[idx["id"]]) == "first-id"


def test_empty_diff_when_upload_subset_of_allowlist(repo_root: Path, tmp_path: Path) -> None:
    tax = repo_root / "doc" / "Taxonomy.csv"
    ref = repo_root / "doc" / "CS_ticket_new_categorizations.xlsx"
    if not tax.is_file() or not ref.is_file():
        pytest.skip("doc artifacts missing")
    allow = load_allowlist(tax, ref)
    existing = next(iter(allow.tuples))
    row = _sample_row(
        Tier1_Segment=existing[0],
        Tier2_Stream=existing[1],
        Tier3_Cat=existing[2],
        Tier4_Type=existing[3],
        Granular_Tech_UI_Type=existing[4],
    )
    upload = tmp_path / "upload.xlsx"
    _write_workbook(upload, [row])
    upload_tuples = extract_workbook_five_tuples(upload)
    assert diff_against_allowlist(upload_tuples, allow) == frozenset()


def _write_workbook_named(path: Path, sheet: str, rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(list(MASTER_COLUMNS))
    for row in rows:
        ws.append([row.get(c, "") for c in MASTER_COLUMNS])
    wb.save(path)


def test_resolve_prefers_scmp_sheet_when_present(tmp_path: Path) -> None:
    xlsx = tmp_path / "both_sheets.xlsx"
    wb = Workbook()
    tickets = wb.active
    tickets.title = "Tickets"
    tickets.append(list(MASTER_COLUMNS))
    tickets.append([_sample_row(id="1").get(c, "") for c in MASTER_COLUMNS])
    scmp = wb.create_sheet("SCMP_Tickets_Master_Categorized")
    scmp.append(list(MASTER_COLUMNS))
    scmp.append([_sample_row(id="2").get(c, "") for c in MASTER_COLUMNS])
    wb.save(xlsx)
    assert resolve_classified_upload_sheet(xlsx) == "SCMP_Tickets_Master_Categorized"


def test_resolve_accepts_portal_tickets_sheet(tmp_path: Path) -> None:
    row = _sample_row(
        id="1",
        Tier1_Segment="B2C",
        Tier2_Stream="Service Task",
        Tier3_Cat="General Support",
        Tier4_Type="Account",
        Granular_Tech_UI_Type="Login",
    )
    xlsx = tmp_path / "portal.xlsx"
    _write_workbook_named(xlsx, "Tickets", [row])
    assert resolve_classified_upload_sheet(xlsx) == "Tickets"
    tuples = extract_classified_workbook_five_tuples(xlsx)
    assert ("B2C", "Service Task", "General Support", "Account", "Login") in tuples


def test_resolve_prefers_largest_valid_sheet(tmp_path: Path) -> None:
    row = _sample_row(
        id="1",
        Tier1_Segment="B2C",
        Tier2_Stream="Service Task",
        Tier3_Cat="General Support",
        Tier4_Type="Account",
        Granular_Tech_UI_Type="Login",
    )
    xlsx = tmp_path / "analyst.xlsx"
    wb = Workbook()
    small = wb.active
    small.title = "Notes"
    small.append(list(MASTER_COLUMNS))
    small.append([row.get(c, "") for c in MASTER_COLUMNS])
    big = wb.create_sheet("AnalystExport")
    big.append(list(MASTER_COLUMNS))
    for i in range(3):
        big.append([dict(row, id=str(i + 1)).get(c, "") for c in MASTER_COLUMNS])
    wb.save(xlsx)
    assert resolve_classified_upload_sheet(xlsx) == "AnalystExport"


def test_resolve_raises_when_no_valid_sheet(tmp_path: Path) -> None:
    xlsx = tmp_path / "empty.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Metadata"
    ws.append(["col1", "col2"])
    ws.append(["a", "b"])
    wb.save(xlsx)
    with pytest.raises(ValueError, match="No sheet with ticket rows"):
        resolve_classified_upload_sheet(xlsx)


def test_merge_uses_classified_source_sheet(tmp_path: Path) -> None:
    novel = ("PortalSeg", "PortalStream", "PortalCat", "PortalType", "PortalGran")
    source = tmp_path / "source.xlsx"
    _write_workbook_named(
        source,
        "Tickets",
        [
            _sample_row(
                id="first-id",
                subject="First",
                Tier1_Segment=novel[0],
                Tier2_Stream=novel[1],
                Tier3_Cat=novel[2],
                Tier4_Type=novel[3],
                Granular_Tech_UI_Type=novel[4],
            )
        ],
    )
    target = tmp_path / "target.xlsx"
    _write_workbook(target, [_sample_row(id="existing")])
    merge_tuples_into_workbook(target, source, frozenset({novel}))
    wb = load_workbook(target, read_only=True, data_only=True)
    ws = wb[SHEET]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    assert len(rows) == 2
    idx = {h: i for i, h in enumerate(MASTER_COLUMNS)}
    assert str(rows[-1][idx["id"]]) == "first-id"


def test_diff_computable_on_repo_workbook(repo_root: Path) -> None:
    tax = repo_root / "doc" / "Taxonomy.csv"
    ref = repo_root / "doc" / "CS_ticket_new_categorizations.xlsx"
    if not tax.is_file() or not ref.is_file():
        pytest.skip("doc artifacts missing")
    allow = load_allowlist(tax, ref)
    upload_tuples = extract_workbook_five_tuples(ref)
    diff_against_allowlist(upload_tuples, allow)
