from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from cs_tickets import portal_app
from cs_tickets.portal_app import app

client = TestClient(app)


def test_health() -> None:
    r = client.get("/health")
    assert r.status_code == 200
    assert r.text == "ok"


def test_index_has_upload_form() -> None:
    r = client.get("/")
    assert r.status_code == 200
    assert "Categorize Support Tickets" in r.text
    assert 'action="/run"' in r.text
    assert "Bad CSAT Rating" in r.text
    assert "/static/classify.js" in r.text
    assert "How Categorization Works (Technical)" in r.text
    assert "readme-doc" not in r.text
    assert "mermaid@10" not in r.text


def test_run_upload_bad_satisfaction_only(repo_root: Path, tmp_path: Path) -> None:
    tax = repo_root / "doc" / "Taxonomy.csv"
    xlsx = repo_root / "doc" / "CS_ticket_new_categorizations.xlsx"
    if not tax.is_file() or not xlsx.is_file():
        pytest.skip("fixtures missing")
    portal_app._RUNS.clear()
    ndjson = (
        '{"url":"https://example.zendesk.com/api/v2/tickets/1.json","id":1,"tags":["app"],'
        '"satisfaction_rating":{"score":"bad"}}\n'
        '{"url":"https://example.zendesk.com/api/v2/tickets/2.json","id":2,"tags":["app"],'
        '"satisfaction_rating":{"score":"good"}}\n'
    )
    r = client.post(
        "/run",
        data={"bad_satisfaction_only": "true"},
        files={"export": ("ratings.ndjson", ndjson.encode(), "application/octet-stream")},
    )
    assert r.status_code == 200
    assert "1 tickets categorized" in r.text
    assert "bad CSAT rating" in r.text


def test_run_upload_ndjson(repo_root: Path) -> None:
    export = repo_root / "tests" / "fixtures" / "five_tickets.ndjson"
    if not export.is_file():
        pytest.skip("fixture missing")
    portal_app._RUNS.clear()
    body = export.read_bytes()
    r = client.post(
        "/run",
        files={"export": ("sample.ndjson", body, "application/octet-stream")},
    )
    assert r.status_code == 200
    assert "tickets categorized" in r.text
    assert "manual review" in r.text
    assert "(TBC)" in r.text
    assert "portal-topnav" in r.text
    assert "Download Excel Workbook" in r.text
    assert "Upload Another File" in r.text
    assert "Run History" in r.text
    assert "drive.google.com/drive/folders/" in r.text
    assert "Results By Category" in r.text
    assert "/static/cs_tickets_theme.css" in r.text
    assert "stats-table" in r.text
    assert "Grand Total" in r.text
    assert "ticket_preview.js" in r.text
    assert "show-ticket-preview-details" in r.text
    assert "show-ticket-preview-tbc-only" in r.text
    assert "preview-col-detail" in r.text
    assert "Why tickets need manual review" in r.text
    assert 'id="classify-ticket-preview-data"' in r.text
    assert "preview-col-detail' hidden" in r.text or 'preview-col-detail" hidden' in r.text
    assert "readme-doc" not in r.text
    run_id = next(iter(portal_app._RUNS))
    assert f'/download/{run_id}"' in r.text


def test_download_workbook_has_tickets_and_tier_tabs(repo_root: Path) -> None:
    export = repo_root / "tests" / "fixtures" / "five_tickets.ndjson"
    if not export.is_file():
        pytest.skip("fixture missing")
    portal_app._RUNS.clear()
    body = export.read_bytes()
    r = client.post(
        "/run",
        files={"export": ("sample.ndjson", body, "application/octet-stream")},
    )
    assert r.status_code == 200
    assert len(portal_app._RUNS) == 1
    run_id = next(iter(portal_app._RUNS))
    d = client.get(f"/download/{run_id}")
    assert d.status_code == 200
    assert "spreadsheetml" in d.headers.get("content-type", "")
    assert d.content[:2] == b"PK"
    wb = load_workbook(BytesIO(d.content), read_only=True, data_only=True)
    assert wb.sheetnames == ["Run metadata", "Tickets", "Tier breakdown"]
    ws = wb["Tier breakdown"]
    tier_rows = list(ws.iter_rows(min_row=1, max_col=5, values_only=True))
    assert tier_rows[0][0] == "Tier1_Segment"
    assert tier_rows[0][4] == "COUNTA of id"
    assert tier_rows[-1][0] == "Grand Total"
    assert isinstance(tier_rows[-1][4], int)
    assert tier_rows[-1][4] == 5


def test_run_upload_invalid_json_returns_400() -> None:
    r = client.post(
        "/run",
        files={"export": ("bad.json", b"not json at all", "application/octet-stream")},
    )
    assert r.status_code == 400
    detail = r.json().get("detail", "")
    assert "Could not parse" in detail or "No JSON ticket" in detail
def test_run_upload_shows_drive_link_when_upload_succeeds(
    repo_root: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    export = repo_root / "tests" / "fixtures" / "five_tickets.ndjson"
    if not export.is_file():
        pytest.skip("fixture missing")
    from cs_tickets.drive_upload import DriveUploadResult

    monkeypatch.setenv("DRIVE_UPLOAD_ENABLED", "true")
    monkeypatch.setenv("GOOGLE_DRIVE_RUNS_FOLDER_ID", "folder-test")
    monkeypatch.setattr(
        "cs_tickets.portal_app.try_upload_workbook",
        lambda _payload, *, filename: (
            DriveUploadResult(
                file_id="abc",
                filename=filename,
                web_view_link="https://drive.google.com/file/d/abc/view",
            ),
            None,
        ),
    )
    body = export.read_bytes()
    r = client.post(
        "/run",
        files={"export": ("sample.ndjson", body, "application/octet-stream")},
    )
    assert r.status_code == 200
    assert "Saved to Google Drive" in r.text
    assert "drive.google.com" in r.text


def test_static_theme_css() -> None:
    r = client.get("/static/cs_tickets_theme.css")
    assert r.status_code == 200
    assert "stats-table" in r.text
    assert r.headers.get("content-type", "").startswith("text/css")


def test_training_link_on_index_when_available(repo_root: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr("cs_tickets.portal_app._repo_root", lambda: repo_root)
    r = client.get("/")
    assert r.status_code == 200
    assert "/learn" in r.text
    assert "Add New Categories" in r.text
    assert "portal-topnav" in r.text


def test_training_redirects_to_learn() -> None:
    r = client.get("/training", follow_redirects=False)
    assert r.status_code == 307
    assert r.headers["location"] == "/learn"


def test_training_post_routes_retired() -> None:
    """Legacy /training POST endpoints were removed; use /learn instead."""
    r = client.post("/training/upload", files={"workbook": ("x.xlsx", b"", "application/octet-stream")})
    assert r.status_code == 404


def test_run_upload_rejects_txt_extension() -> None:
    r = client.post(
        "/run",
        files={"export": ("export.txt", b'{"id": 1}\n', "text/plain")},
    )
    assert r.status_code == 400
    assert "Export file must be one of" in r.json().get("detail", "")


